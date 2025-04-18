import streamlit as st
from database import setup_database, view_record_by_id, view_all_records, delete_all_records
from form_handler import create_registration_form
from openpyxl import Workbook, load_workbook
import os
import pandas as pd

# Carregar variáveis de ambiente do arquivo .env

def create_excel_with_model(data):
    # Carregar o modelo de Excel
    wb = load_workbook('/Users/Samsung/Desktop/bb/project/modelo_devolucao.xlsx')
    ws = wb['protocolo']

    # Preencher as células especificadas
    for entry in data:
        rota, motorista, transportadora, pedido, remessa, nf, data_registro = entry[1], entry[2], entry[3], entry[4], entry[5], entry[6], entry[8]

        def split_info(info):
            separators = ['/','-',',','.',':',';','_']
            for separator in separators:
                if separator in info:
                    return info.split(separator)
            return [info]

        pedido_list = split_info(pedido)
        remessa_list = split_info(remessa)
        nf_list = split_info(nf)

        # Preencher a primeira parte
        ws['A12'] = rota
        for idx, value in enumerate(pedido_list, start=12):
            ws.cell(row=idx, column=2).value = value
        for idx, value in enumerate(remessa_list, start=12):
            ws.cell(row=idx, column=3).value = value
        for idx, value in enumerate(nf_list, start=12):
            ws.cell(row=idx, column=4).value = value
        ws['B25'] = motorista
        ws['C27'] = transportadora
        ws['B31'] = data_registro

        # Preencher a segunda parte
        ws['I12'] = rota
        for idx, value in enumerate(pedido_list, start=12):
            ws.cell(row=idx, column=10).value = value
        for idx, value in enumerate(remessa_list, start=12):
            ws.cell(row=idx, column=11).value = value
        for idx, value in enumerate(nf_list, start=12):
            ws.cell(row=idx, column=12).value = value
        ws['J25'] = motorista
        ws['K27'] = transportadora
        ws['I31'] = data_registro

    file_path = 'Protocolo_de_Entregas_Atualizado.xlsx'
    wb.save(file_path)
    return file_path

def export_all_to_excel(data):
    # Criar uma nova planilha
    wb = Workbook()
    ws = wb.active
    ws.title = "Registros Exportados"

    # Adicionar os cabeçalhos personalizados
    headers = ["ID", "Rota", "Motorista", "Transportadora", "Pedido", "Remessa", "NF", "Motivo da Devolução", "Data"]
    ws.append(headers)

    # Adicionar os dados
    for row in data:
        ws.append(row)

    file_path = 'Registros_Exportados.xlsx'
    wb.save(file_path)
    return file_path

def main():
    st.set_page_config(
        page_title="Protocolo de Devolução/Reentrega",
        page_icon="📋",
        layout="wide"
    )
    
    # Initialize database
    setup_database()
    
    st.title("📋 Protocolo de Devolução/Reentrega")
    
    # Create tabs for better organization
    tab1, tab2, tab3, tab4 = st.tabs(["Registrar Protocolo", "Consultar por ID", "Todos os Registros", "Administração"])
    
    with tab1:
        st.header("Registro de Protocolo")
        create_registration_form()
    
    with tab2:
        st.header("Consultar Protocolo por ID")
        id_input = st.text_input("Digite o ID do protocolo:", key="id_lookup")
        
        if st.button("Consultar", key="btn_consultar"):
            if id_input:
                try:
                    id_value = int(id_input)
                    result = view_record_by_id(id_value)
                    
                    if result:
                        st.success(f"Protocolo #{result[0]} encontrado")
                        
                        # Display in a nice card format
                        col1, col2 = st.columns(2)
                        with col1:
                            st.markdown("### Informações da Rota")
                            st.write(f"**Rota:** {result[1]}")
                            st.write(f"**Motorista:** {result[2]}")
                            st.write(f"**Transportadora:** {result[3]}")
                            
                        with col2:
                            st.markdown("### Detalhes do Pedido")
                            st.write(f"**Pedido:** {result[4]}")
                            st.write(f"**Remessa:** {result[5]}")
                            st.write(f"**Nota Fiscal:** {result[6]}")
                            st.write(f"**Motivo:** {result[7]}")
                            st.write(f"**Data:** {result[8]}")
                        
                        data = [result]
                        file_path = create_excel_with_model(data)
                        
                        with open(file_path, "rb") as file:
                            st.download_button(
                                label="📥 Exportar registro para Excel",
                                data=file,
                                file_name="Protocolo_de_Entregas_Atualizado.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                    else:
                        st.error("Nenhum registro encontrado para o ID fornecido.")
                except ValueError:
                    st.error("Por favor, insira um ID válido.")
            else:
                st.info("Digite um ID para consultar.")
    
    with tab3:
        st.header("Todos os Registros")
        if st.button("Mostrar Todos", key="btn_mostrar_todos"):
            records = view_all_records()
            if records:
                st.success(f"{len(records)} registros encontrados")
                
                # Criar DataFrame com cabeçalhos personalizados
                df = pd.DataFrame(
                    records,
                    columns=["ID", "Rota", "Motorista", "Transportadora", "Pedido", "Remessa", "NF", "Motivo da Devolução", "Data"]
                )
                
                # Exibir a tabela no Streamlit
                st.dataframe(df, hide_index=True)
                
                file_path = export_all_to_excel(records)
                
                with open(file_path, "rb") as file:
                    st.download_button(
                        label="📥 Exportar Todos para Excel",
                        data=file,
                        file_name="Registros_Exportados.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            else:
                st.info("Nenhum registro encontrado no banco de dados.")
    
    with tab4:
        st.header("Administração")
        with st.expander("Excluir Todos os Registros"):
            st.warning("⚠️ Esta ação não pode ser desfeita!")
            password_input = st.text_input("Digite a senha para excluir todos os registros:", type="password", key="delete_password")
            
            if st.button("Excluir Todos os Registros", key="btn_delete_all"):
                if delete_all_records(password_input):
                    st.success("Todos os registros foram excluídos com sucesso.")
                else:
                    st.error("Senha incorreta. Não foi possível excluir os registros.")

if __name__ == "__main__":
    main()
