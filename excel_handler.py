import streamlit as st
from openpyxl import Workbook, load_workbook
import os
import tempfile

def split_info(info):
    """Split information using common separators"""
    if not info:
        return []
        
    separators = ['/','-',',','.',':',';','_']
    for separator in separators:
        if separator in info:
            return info.split(separator)
    return [info]

def create_protocol_excel(data):
    """Create an Excel file for a single protocol"""
    try:
        # Create a temporary file
        temp_dir = tempfile.gettempdir()
        output_file = os.path.join(temp_dir, 'Protocolo_de_Entregas.xlsx')
        
        # Try to locate the template file in Streamlit's static files or create a new workbook
        template_path = None
        try:
            template_path = st.secrets["excel"]["template_path"]
            wb = load_workbook(template_path)
            ws = wb['protocolo']
        except (KeyError, FileNotFoundError):
            # If template not found, create a basic Excel file
            wb = Workbook()
            ws = wb.active
            ws.title = "protocolo"
            # Create headers
            headers = ['Rota', 'Pedido', 'Remessa', 'Nota Fiscal', 'Motorista', 'Transportadora', 'Data']
            for col, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col).value = header
        
        # If we have the template, fill it according to the template format
        if template_path:
            for entry in data:
                rota, motorista, transportadora = entry[1], entry[2], entry[3]
                pedido, remessa, nf = entry[4], entry[5], entry[6]
                data_registro = entry[8]
                
                pedido_list = split_info(pedido)
                remessa_list = split_info(remessa)
                nf_list = split_info(nf)
                
                # Preencher a primeira parte
                ws['A12'] = rota
                for idx, value in enumerate(pedido_list, start=12):
                    ws.cell(row=idx, column=2).value = value
                for idx, value in enumerate(remessa_list, start=12):
                    ws.cell(row=idx, column=3).value = value
                for idx, value in enumerate(nf_list, start=12):
                    ws.cell(row=idx, column=4).value = value
                ws['B25'] = motorista
                ws['C27'] = transportadora
                ws['B31'] = data_registro
                
                # Preencher a segunda parte
                ws['I12'] = rota
                for idx, value in enumerate(pedido_list, start=12):
                    ws.cell(row=idx, column=10).value = value  # J12 é a coluna 10
                for idx, value in enumerate(remessa_list, start=12):
                    ws.cell(row=idx, column=11).value = value  # K12 é a coluna 11
                for idx, value in enumerate(nf_list, start=12):
                    ws.cell(row=idx, column=12).value = value  # L12 é a coluna 12
                ws['J25'] = motorista
                ws['K27'] = transportadora
                ws['I31'] = data_registro
        else:
            # Basic Excel formatting if template not found
            row_num = 2
            for entry in data:
                ws.cell(row=row_num, column=1).value = entry[1]  # Rota
                ws.cell(row=row_num, column=2).value = entry[4]  # Pedido
                ws.cell(row=row_num, column=3).value = entry[5]  # Remessa
                ws.cell(row=row_num, column=4).value = entry[6]  # Nota Fiscal
                ws.cell(row=row_num, column=5).value = entry[2]  # Motorista
                ws.cell(row=row_num, column=6).value = entry[3]  # Transportadora
                ws.cell(row=row_num, column=7).value = entry[8]  # Data
                row_num += 1
                
        # Save the workbook
        wb.save(output_file)
        return output_file
    
    except Exception as e:
        st.error(f"Erro ao criar arquivo Excel: {e}")
        # Create a basic fallback Excel file
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Erro ao gerar relatório"
        ws['A2'] = str(e)
        
        fallback_file = os.path.join(tempfile.gettempdir(), 'Erro_Protocolo.xlsx')
        wb.save(fallback_file)
        return fallback_file

def export_all_records_excel(data):
    """Export all records to an Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Registros"
    
    # Add headers
    headers = ['ID', 'Rota', 'Motorista', 'Transportadora', 'Pedido', 'Remessa', 'Nota Fiscal', 'Motivo', 'Data']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = header
    
    # Add data
    for row_idx, row in enumerate(data, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx).value = value
    
    # Auto adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save to a temporary file
    temp_dir = tempfile.gettempdir()
    file_path = os.path.join(temp_dir, 'Registros_Exportados.xlsx')
    wb.save(file_path)
    return file_path